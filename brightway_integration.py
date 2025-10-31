"""
Brightway 2.5 Integration for Sustain 4.0 BioEngine
Handles LCI data import, validation, and database creation
"""

import pandas as pd
from typing import Dict, List, Tuple, Optional
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import plotly.graph_objects as go

# Brightway imports - optional, only needed when actually creating databases
try:
    import bw2data as bd
    import bw2io as bi
    BRIGHTWAY_AVAILABLE = True
except ImportError:
    BRIGHTWAY_AVAILABLE = False


class SustainExcelImporter:
    """
    Import LCI data from Sustain 4.0 Excel template into Brightway
    """
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.metadata = {}
        self.activities = []
        self.exchanges = []
        self.flow_mapping = {}
        self.validation_errors = []
        self.warnings = []
        
    def parse_excel(self) -> bool:
        """Parse Excel file and validate structure"""
        
        try:
            # Read all sheets
            xl_file = pd.ExcelFile(self.excel_path)
            
            # Check required sheets
            required_sheets = ['Project Metadata', 'Process Activities', 'Exchanges']
            for sheet in required_sheets:
                if sheet not in xl_file.sheet_names:
                    self.validation_errors.append(f"❌ Missing required sheet: {sheet}")
                    return False
            
            # Parse each sheet
            self.metadata = self._parse_metadata(xl_file)
            self.activities = self._parse_activities(xl_file)
            self.exchanges = self._parse_exchanges(xl_file)
            
            # Optional: Flow mapping
            if 'Biosphere Flows Mapping' in xl_file.sheet_names:
                self.flow_mapping = self._parse_flow_mapping(xl_file)
            
            # Validate data
            self._validate_data()
            
            return len(self.validation_errors) == 0
            
        except Exception as e:
            self.validation_errors.append(f"❌ Error parsing Excel: {str(e)}")
            return False
    
    def _parse_metadata(self, xl_file) -> dict:
        """Parse project metadata sheet"""
        
        df = pd.read_excel(xl_file, sheet_name='Project Metadata', header=None)
        
        metadata = {}
        for _, row in df.iterrows():
            if pd.notna(row[0]) and pd.notna(row[1]):
                metadata[str(row[0]).strip()] = str(row[1]).strip()
        
        return metadata
    
    def _parse_activities(self, xl_file) -> List[dict]:
        """Parse process activities sheet"""
        
        df = pd.read_excel(xl_file, sheet_name='Process Activities')
        
        activities = []
        for _, row in df.iterrows():
            if pd.notna(row['Activity Code']):
                activities.append({
                    'code': str(row['Activity Code']).strip(),
                    'name': str(row['Activity Name']).strip(),
                    'unit': str(row['Unit']).strip(),
                    'location': str(row['Location']).strip(),
                    'reference_production': float(row['Reference Production'])
                })
        
        return activities
    
    def _parse_exchanges(self, xl_file) -> List[dict]:
        """Parse exchanges sheet"""
        
        df = pd.read_excel(xl_file, sheet_name='Exchanges')
        
        exchanges = []
        for _, row in df.iterrows():
            if pd.notna(row['Activity Code']):
                exc = {
                    'activity_code': str(row['Activity Code']).strip(),
                    'type': str(row['Exchange Type']).strip().lower(),
                    'flow_name': str(row['Flow Name']).strip(),
                    'amount': float(row['Amount']),
                    'unit': str(row['Unit']).strip(),
                    'category': str(row['Category']).strip() if pd.notna(row['Category']) else None
                }
                
                # Parse uncertainty if provided
                if 'Uncertainty' in df.columns and pd.notna(row['Uncertainty']):
                    uncertainty_str = str(row['Uncertainty'])
                    if '±' in uncertainty_str:
                        try:
                            exc['uncertainty'] = float(uncertainty_str.replace('±', '').strip())
                        except:
                            pass
                
                exchanges.append(exc)
        
        return exchanges
    
    def _parse_flow_mapping(self, xl_file) -> dict:
        """Parse biosphere flow mapping sheet"""
        
        df = pd.read_excel(xl_file, sheet_name='Biosphere Flows Mapping')
        
        mapping = {}
        for _, row in df.iterrows():
            if pd.notna(row['User Flow Name']) and pd.notna(row['Biosphere3 Flow Name (Brightway)']):
                user_name = str(row['User Flow Name']).strip()
                biosphere_name = str(row['Biosphere3 Flow Name (Brightway)']).strip()
                mapping[user_name] = biosphere_name
        
        return mapping
    
    def _validate_data(self):
        """Validate data consistency"""
        
        # 1. Check all activity codes in exchanges exist in activities
        activity_codes = {act['code'] for act in self.activities}
        exchange_codes = {exc['activity_code'] for exc in self.exchanges}
        
        orphan_codes = exchange_codes - activity_codes
        if orphan_codes:
            self.validation_errors.append(
                f"❌ Exchanges reference non-existent activities: {orphan_codes}"
            )
        
        # 2. Check each activity has a production exchange
        for activity in self.activities:
            has_production = any(
                exc['activity_code'] == activity['code'] and exc['type'] == 'production'
                for exc in self.exchanges
            )
            if not has_production:
                self.validation_errors.append(
                    f"❌ Activity '{activity['name']}' has no production exchange"
                )
        
        # 3. Check for negative amounts
        for exc in self.exchanges:
            if exc['amount'] < 0:
                self.warnings.append(
                    f"⚠️ Negative amount in {exc['activity_code']}: {exc['flow_name']}"
                )
        
        # 4. Basic mass balance check (optional, only warning)
        for activity in self.activities:
            inputs = [exc for exc in self.exchanges 
                     if exc['activity_code'] == activity['code'] and exc['type'] == 'input']
            outputs = [exc for exc in self.exchanges 
                      if exc['activity_code'] == activity['code'] and 
                      (exc['type'] == 'production' or exc['type'] == 'emission')]
            
            # Simplified check (only for mass units)
            mass_units = ['kg', 'g', 't', 'ton']
            input_mass = sum(exc['amount'] for exc in inputs 
                           if exc['unit'] in mass_units)
            output_mass = sum(exc['amount'] for exc in outputs 
                            if exc['unit'] in mass_units)
            
            if input_mass > 0 and output_mass > 0:
                balance_ratio = output_mass / input_mass
                if balance_ratio < 0.3 or balance_ratio > 1.2:
                    self.warnings.append(
                        f"⚠️ Suspicious mass balance in '{activity['name']}': "
                        f"Input={input_mass:.2f}kg, Output={output_mass:.2f}kg (ratio: {balance_ratio:.2f})"
                    )
    
    def link_biosphere_flows(self) -> Dict[str, Tuple[str, int]]:
        """
        Link user flow names to biosphere3 database
        
        Returns:
            dict: {user_flow_name: (biosphere_db, biosphere_code)}
        """
        
        try:
            biosphere_db = bd.Database('biosphere3')
        except:
            self.warnings.append("⚠️ Biosphere3 database not found. Flows will not be linked.")
            return {}
        
        linked_flows = {}
        unlinked_flows = []
        
        # Get unique emission/resource flow names
        biosphere_exchanges = [
            exc for exc in self.exchanges 
            if exc['type'] in ['emission', 'resource']
        ]
        
        unique_flows = {exc['flow_name'] for exc in biosphere_exchanges}
        
        for flow_name in unique_flows:
            # Check if user provided mapping
            if flow_name in self.flow_mapping:
                search_term = self.flow_mapping[flow_name]
            else:
                search_term = flow_name
            
            # Search biosphere3
            try:
                results = biosphere_db.search(search_term)
                
                if results:
                    # Take best match (first result)
                    best_match = results[0]
                    linked_flows[flow_name] = ('biosphere3', best_match['code'])
                else:
                    unlinked_flows.append(flow_name)
            except:
                unlinked_flows.append(flow_name)
        
        if unlinked_flows:
            self.warnings.append(
                f"⚠️ Could not automatically link flows: {unlinked_flows}. "
                f"They will be created as generic flows."
            )
        
        return linked_flows
    
    def create_brightway_database(self, db_name: str, project_name: str = None) -> str:
        """
        Create Brightway database from parsed data
        
        Args:
            db_name: Name for the new database
            project_name: Brightway project name (optional)
        
        Returns:
            Database name if successful
        """
        
        if self.validation_errors:
            raise ValueError(f"Cannot create database with validation errors: {self.validation_errors}")
        
        # Set project if specified
        if project_name:
            try:
                bd.projects.set_current(project_name)
            except:
                # If project doesn't exist, it will be created
                bd.projects.set_current(project_name)
        
        # Link biosphere flows
        linked_flows = self.link_biosphere_flows()
        
        # Build database structure
        db_data = {}
        
        for activity in self.activities:
            activity_key = (db_name, activity['code'])
            
            # Get exchanges for this activity
            activity_exchanges = [
                exc for exc in self.exchanges 
                if exc['activity_code'] == activity['code']
            ]
            
            # Convert exchanges to Brightway format
            bw_exchanges = []
            
            for exc in activity_exchanges:
                bw_exc = {
                    'amount': exc['amount'],
                    'unit': exc['unit']
                }
                
                # Determine input key
                if exc['type'] == 'production':
                    bw_exc['input'] = activity_key
                    bw_exc['type'] = 'production'
                
                elif exc['type'] == 'input':
                    # Try to find if it's another activity in this database
                    matching_activity = next(
                        (act for act in self.activities if act['name'] == exc['flow_name']),
                        None
                    )
                    
                    if matching_activity:
                        # Internal link
                        bw_exc['input'] = (db_name, matching_activity['code'])
                        bw_exc['type'] = 'technosphere'
                    else:
                        # External input - create as generic technosphere
                        bw_exc['input'] = (db_name, f"generic_{exc['flow_name']}")
                        bw_exc['type'] = 'technosphere'
                        bw_exc['name'] = exc['flow_name']
                
                elif exc['type'] in ['emission', 'resource']:
                    # Link to biosphere
                    if exc['flow_name'] in linked_flows:
                        bw_exc['input'] = linked_flows[exc['flow_name']]
                    else:
                        # Unlinked - create as generic biosphere
                        bw_exc['input'] = (db_name, f"bio_{exc['flow_name']}")
                        bw_exc['name'] = exc['flow_name']
                    
                    bw_exc['type'] = 'biosphere'
                    if exc['category']:
                        bw_exc['categories'] = (exc['category'],)
                
                # Add uncertainty if available
                if 'uncertainty' in exc:
                    bw_exc['uncertainty type'] = 3  # Normal distribution
                    bw_exc['loc'] = exc['amount']
                    bw_exc['scale'] = exc['uncertainty']
                
                bw_exchanges.append(bw_exc)
            
            # Create activity
            db_data[activity_key] = {
                'name': activity['name'],
                'unit': activity['unit'],
                'location': activity['location'],
                'type': 'process',
                'exchanges': bw_exchanges,
                'production amount': activity['reference_production']
            }
        
        # Write database
        if db_name in bd.databases:
            del bd.databases[db_name]
        
        db = bd.Database(db_name)
        db.write(db_data)
        
        return db_name


def generate_excel_template(project_data: dict) -> BytesIO:
    """Generate Excel template for user to fill with LCI data"""
    
    wb = Workbook()
    
    # Sheet 1: Metadata
    ws1 = wb.active
    ws1.title = "Project Metadata"
    
    ws1.append(["Project Name", project_data.get('name', '')])
    ws1.append(["Functional Unit", f"1 {project_data.get('reference_flow_unit', 'kg')}"])
    ws1.append(["Location", project_data.get('region', 'BR')])
    ws1.append(["Scale", project_data.get('scale', 'pilot')])
    ws1.append(["System Boundaries", project_data.get('system_boundaries', 'gate-to-gate')])
    
    # Sheet 2: Activities
    ws2 = wb.create_sheet("Process Activities")
    ws2.append(["Activity Code", "Activity Name", "Unit", "Location", "Reference Production"])
    ws2.append(["PROC_001", "Your Process 1", "kg", project_data.get('region', 'BR'), "1.0"])
    ws2.append(["PROC_002", "Your Process 2", "L", project_data.get('region', 'BR'), "1.0"])
    
    # Sheet 3: Exchanges
    ws3 = wb.create_sheet("Exchanges")
    headers = ["Activity Code", "Exchange Type", "Flow Name", "Amount", "Unit", "Category", "Uncertainty"]
    ws3.append(headers)
    
    # Add examples
    ws3.append(["PROC_001", "production", "Product 1", "1.0", "kg", "", ""])
    ws3.append(["PROC_001", "input", "Raw material", "2.0", "kg", "material", "±0.1"])
    ws3.append(["PROC_001", "input", "Electricity", "1.5", "kWh", "energy", "±0.1"])
    ws3.append(["PROC_001", "emission", "CO2", "0.5", "kg", "air", "±0.05"])
    
    # Sheet 4: Flow Mapping (optional)
    ws4 = wb.create_sheet("Biosphere Flows Mapping")
    ws4.append(["User Flow Name", "Biosphere3 Flow Name (Brightway)"])
    ws4.append(["CO2", "Carbon dioxide, fossil"])
    ws4.append(["CH4", "Methane, fossil"])
    ws4.append(["Water", "Water, unspecified natural origin"])
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for ws in [ws1, ws2, ws3, ws4]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def get_example_lci_file() -> BytesIO:
    """Generate example LCI file with realistic data"""
    
    wb = Workbook()
    
    # Sheet 1: Metadata
    ws1 = wb.active
    ws1.title = "Project Metadata"
    ws1.append(["Project Name", "Bioethanol from Sugarcane - Example"])
    ws1.append(["Functional Unit", "1 L"])
    ws1.append(["Location", "BR"])
    ws1.append(["Scale", "pilot"])
    ws1.append(["System Boundaries", "gate-to-gate"])
    
    # Sheet 2: Activities
    ws2 = wb.create_sheet("Process Activities")
    ws2.append(["Activity Code", "Activity Name", "Unit", "Location", "Reference Production"])
    ws2.append(["FERM_01", "Fermentation", "kg", "BR", "1.0"])
    ws2.append(["DIST_01", "Distillation", "L", "BR", "1.0"])
    
    # Sheet 3: Exchanges
    ws3 = wb.create_sheet("Exchanges")
    headers = ["Activity Code", "Exchange Type", "Flow Name", "Amount", "Unit", "Category", "Uncertainty"]
    ws3.append(headers)
    
    # Fermentation exchanges
    ws3.append(["FERM_01", "production", "Fermented mass", "1.0", "kg", "", ""])
    ws3.append(["FERM_01", "input", "Sugarcane juice", "1.8", "kg", "material", "±0.1"])
    ws3.append(["FERM_01", "input", "Yeast", "0.05", "kg", "material", "±0.005"])
    ws3.append(["FERM_01", "input", "Water", "5.0", "kg", "material", "±0.25"])
    ws3.append(["FERM_01", "input", "Electricity", "2.5", "kWh", "energy", "±0.15"])
    ws3.append(["FERM_01", "emission", "CO2, biogenic", "0.9", "kg", "air", "±0.05"])
    ws3.append(["FERM_01", "emission", "Wastewater", "4.5", "kg", "water", "±0.3"])
    
    # Distillation exchanges
    ws3.append(["DIST_01", "production", "Crude ethanol", "1.0", "L", "", ""])
    ws3.append(["DIST_01", "input", "Fermented mass", "1.2", "kg", "material", "±0.08"])
    ws3.append(["DIST_01", "input", "Heat (steam)", "15.0", "MJ", "energy", "±1.0"])
    ws3.append(["DIST_01", "emission", "Ethanol vapor", "0.02", "kg", "air", "±0.003"])
    
    # Sheet 4: Flow Mapping
    ws4 = wb.create_sheet("Biosphere Flows Mapping")
    ws4.append(["User Flow Name", "Biosphere3 Flow Name (Brightway)"])
    ws4.append(["CO2, biogenic", "Carbon dioxide, non-fossil"])
    ws4.append(["Wastewater", "Water, unspecified natural origin"])
    ws4.append(["Ethanol vapor", "Ethanol"])
    
    # Style headers
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for ws in [ws1, ws2, ws3, ws4]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def generate_process_network_diagram(activities: List[dict], exchanges: List[dict]):
    """Generate interactive network diagram of process flow"""
    
    try:
        import networkx as nx
    except ImportError:
        # If networkx is not installed, return a simple message
        fig = go.Figure()
        fig.add_annotation(
            text="Network diagram requires 'networkx' package.<br>Install with: pip install networkx",
            xref="paper", yref="paper",
            x=0.5, y=0.5,
            showarrow=False,
            font=dict(size=14, color="gray")
        )
        fig.update_layout(
            xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
            yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
            height=400
        )
        return fig
    
    # Create directed graph
    G = nx.DiGraph()
    
    # Add nodes
    for activity in activities:
        G.add_node(activity['code'], label=activity['name'], unit=activity['unit'])
    
    # Add edges (only internal connections)
    for exc in exchanges:
        if exc['type'] == 'input':
            # Find source activity
            source = next(
                (act['code'] for act in activities if act['name'] == exc['flow_name']),
                None
            )
            if source and source != exc['activity_code']:
                G.add_edge(source, exc['activity_code'], 
                          weight=exc['amount'], 
                          label=f"{exc['amount']:.2f} {exc['unit']}")
    
    # Layout
    if len(G.nodes()) > 0:
        pos = nx.spring_layout(G, k=2, iterations=50)
    else:
        pos = {}
    
    # Create plotly figure
    edge_trace = go.Scatter(
        x=[], y=[],
        line=dict(width=2, color='#888'),
        hoverinfo='text',
        mode='lines',
        text=[]
    )
    
    edge_annotations = []
    
    for edge in G.edges(data=True):
        x0, y0 = pos[edge[0]]
        x1, y1 = pos[edge[1]]
        edge_trace['x'] += tuple([x0, x1, None])
        edge_trace['y'] += tuple([y0, y1, None])
        
        # Add edge label
        edge_annotations.append(
            dict(
                x=(x0 + x1) / 2,
                y=(y0 + y1) / 2,
                text=edge[2].get('label', ''),
                showarrow=False,
                font=dict(size=10, color='#666')
            )
        )
    
    node_trace = go.Scatter(
        x=[], y=[],
        text=[],
        mode='markers+text',
        textposition="top center",
        hoverinfo='text',
        marker=dict(
            size=30,
            color='#4472C4',
            line=dict(width=2, color='white')
        ),
        textfont=dict(size=12, color='#2c3e50')
    )
    
    for node in G.nodes(data=True):
        x, y = pos[node[0]]
        node_trace['x'] += tuple([x])
        node_trace['y'] += tuple([y])
        node_trace['text'] += tuple([node[1]['label']])
    
    fig = go.Figure(data=[edge_trace, node_trace],
                   layout=go.Layout(
                       title="Process Network",
                       showlegend=False,
                       hovermode='closest',
                       annotations=edge_annotations,
                       xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                       yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                       plot_bgcolor='rgba(0,0,0,0)',
                       paper_bgcolor='rgba(0,0,0,0)',
                       height=400
                   ))
    
    return fig
